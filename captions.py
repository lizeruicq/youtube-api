
import google_auth_oauthlib.flow
import googleapiclient.discovery
import googleapiclient.errors
import tokensupport
from googleapiclient.http import MediaFileUpload
from openpyxl import Workbook,load_workbook
wb = load_workbook('data.xlsx')
ws = wb['Sheet1']

partlist = []
filelist = []
diclist=[]
# key = ws.cell(2,1).value
def reaadtable():
    for i in range(2,ws.max_row+1):
        partlist.append(ws.cell(i,1).value)
    # print("partlist:"+str(partlist))

    for i in range(2,ws.max_row+1):
        filelist.append(ws.cell(i,6).value)
    # print("filelist:"+str(filelist))

    # bigdic = {}
    for i in range(2,ws.max_row+1):

        dic = {}
        for j in range(2,6):
            skey = ws.cell(1,j).value
            svalue = ws.cell(i,j).value
            dic[skey] = svalue
        # bigdic[key] = dic
        diclist.append(dic)
    # print("dic:"+str(diclist))
# part1 = 'snippet'
# body1 = {
# 'snippet': {'language': 'en','name': '英语字幕','videoId': 'BS_1MWu6jTA','isDraft': False}
# }
# file1 = '5.4PV_EN.srt'
#
# part2 = "snippet"
# body2 = {
# "snippet": {"language": "zh-tw","name": "繁中字幕","videoId": "BS_1MWu6jTA","isDraft": False}
# }
# file2 = "5.4PV_TW.srt"
#
# parts = [part1]
# bodys = [body1]
# files = [file1]
def main():
    reaadtable()
    for i in range(0,len(partlist)):
        a = partlist[i]
        b = {ws.cell(2,1).value:diclist[i]}
        c = filelist[i]
        # print(a)
        # print(b)
        # print(c)
        print("正在添加"+str(c))
        upload_captions(a,b,c)

def upload_captions(a,b,c):
    # scopes = ["https://www.googleapis.com/auth/youtube.force-ssl"]
    # os.environ["OAUTHLIB_INSECURE_TRANSPORT"] = "1"
    # api_service_name = "youtube"
    # api_version = "v3"
    # client_secrets_file = "123.json"
    # flow = google_auth_oauthlib.flow.InstalledAppFlow.from_client_secrets_file(
    #     client_secrets_file, scopes)
    # credentials = flow.run_console()
    # youtube = googleapiclient.discovery.build(
    #     api_service_name, api_version, credentials=credentials)
    youtube = tokensupport.create_service("cred.json",
                             ["https://www.googleapis.com/auth/youtube.force-ssl"])
    if not youtube: return

    request = youtube.captions().insert(
        part=a,
        body=b,
        media_body=MediaFileUpload(c)
    )

    response = request.execute()


    print(response)

if __name__ == "__main__":
    main()